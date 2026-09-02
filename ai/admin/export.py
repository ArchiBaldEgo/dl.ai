"""XLSX-экспорт результатов batch-прогона ARM (матрица задача×модель).

Единый серверный генератор для журнала запросов (запись batch-лога) и для
страницы /arm/solve/ (живой/завершённый прогон по run_id) — раньше формат
существовал в двух зеркалах (серверный CSV в logs.py и клиентский
downloadResultsCsv в _ai_batch_results.html).

Лист 1 «Сводка»: матрица задача×модель с сокращёнными названиями моделей
(short_model_titles) + сводная таблица по моделям (% решено, среднее время).
Лист 2 «Расшифровка моделей»: сокращение → полное название — читаемость
любого кода гарантирована даже при агрессивном сокращении.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..arm_runner import _per_bucket
from ..model_clients.registry import short_model_titles

# Границы автоширины (в символах Excel): короткие колонки не схлопываются,
# длинные названия задач/моделей не растягивают лист без предела.
_MIN_COL_WIDTH = 8
_MAX_COL_WIDTH = 60

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _verdict_mark(verdict):
    """'+' solved / '-' failed / '?' прочее — прежняя CSV-нотация."""
    if verdict == "solved":
        return "+"
    if verdict == "failed":
        return "-"
    return "?"


def _autofit_columns(ws, header, rows):
    """Автоширина: максимум по длине значений колонки (заголовок входит),
    с разумными границами."""
    all_rows = [header] + rows
    for col_idx in range(1, len(header) + 1):
        width = max((len(str(row[col_idx - 1])) for row in all_rows), default=0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(width + 2, _MIN_COL_WIDTH), _MAX_COL_WIDTH,
        )


def _arm_results_matrix(results, title_map=None):
    """Матрица задача×модель: (header, rows).

    header = ["Node ID", "Задача", <сокращённые названия моделей>]; в ячейках
    '+/-/?'. Порядок задач/моделей — как в списке результатов (первое
    появление). ``title_map`` — мапа «полное название → сокращение»; без неё
    используются полные названия.
    """
    task_map, task_order = {}, []
    model_map, model_order = {}, []
    matrix = {}
    for r in results:
        tkey = str(r.get("task_node_id") or "")
        mkey = str(r.get("model_key") or r.get("model_title") or "")
        if tkey not in task_map:
            task_map[tkey] = r.get("task_name") or ""
            task_order.append(tkey)
        if mkey not in model_map:
            model_map[mkey] = r.get("model_title") or mkey
            model_order.append(mkey)
        matrix.setdefault(tkey, {})[mkey] = _verdict_mark(r.get("verdict"))

    header = ["Node ID", "Задача"] + [
        (title_map or {}).get(model_map[m], model_map[m]) for m in model_order
    ]
    rows = []
    for tkey in task_order:
        # Node ID пишем числом — текстовая ячейка ломает сортировку в Excel.
        node_cell = int(tkey) if str(tkey).isdigit() else tkey
        row = [node_cell, task_map[tkey]]
        row.extend(matrix.get(tkey, {}).get(m, "") for m in model_order)
        rows.append(row)
    return header, rows


def _per_model_summary_rows(results, title_map=None):
    """Сводная таблица по моделям под матрицей: переиспользует
    arm_runner._per_bucket (DRY — та же семантика skipped/avg, что и в UI)."""
    per_model = _per_bucket(
        results,
        key_fn=lambda r: r.get("model_key") or r.get("model_title"),
        label_fn=lambda r: r.get("model_title") or r.get("model_key") or "?",
    )
    header = ["Модель", "% решено", "Среднее время, сек", "Токены", "Решено/всего"]
    rows = []
    for bucket in per_model:
        rows.append([
            (title_map or {}).get(bucket["label"], bucket["label"]),
            bucket["percent_solved"],
            bucket["avg_duration"],
            bucket["tokens"],
            f'{bucket["solved"]}/{bucket["total"]}',
        ])
    return header, rows


def _model_legend_rows(results, title_map=None):
    """Лист-расшифровка: [сокращение, полное название, провайдер].

    Провайдер — первый токен полного названия (Ollama/Web/…)."""
    titles = []
    seen = set()
    for r in results:
        title = r.get("model_title") or r.get("model_key") or ""
        if title and title not in seen:
            seen.add(title)
            titles.append(title)
    provider = lambda title: (str(title).strip().split() or ["?"])[0]  # noqa: E731
    return (
        ["Сокращение", "Полное название", "Провайдер"],
        [
            [(title_map or {}).get(t, t), t, provider(t)]
            for t in titles
        ],
    )


def build_arm_results_xlsx(results) -> bytes:
    """Собрать .xlsx: лист «Сводка» (матрица + сводка по моделям) и лист
    «Расшифровка моделей» (сокращение → полное название)."""
    title_map = short_model_titles(
        [r.get("model_title") or r.get("model_key") or ""
         for r in results if (r.get("model_title") or r.get("model_key"))]
    )

    wb = Workbook()

    # Лист 1 — сводка: матрица задача×модель + таблица по моделям.
    ws = wb.active
    ws.title = "Сводка"
    matrix_header, matrix_rows = _arm_results_matrix(results, title_map)
    for row in [matrix_header] + matrix_rows:
        ws.append(row)
    summary_header, summary_rows = _per_model_summary_rows(results, title_map)
    ws.append([])
    ws.append(summary_header)
    for row in summary_rows:
        ws.append(row)
    _autofit_columns(ws, matrix_header, matrix_rows)

    # Лист 2 — расшифровка сокращений.
    ws_legend = wb.create_sheet("Расшифровка моделей")
    legend_header, legend_rows = _model_legend_rows(results, title_map)
    for row in [legend_header] + legend_rows:
        ws_legend.append(row)
    _autofit_columns(ws_legend, legend_header, legend_rows)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()